from fastapi import FastAPI, APIRouter, HTTPException, UploadFile, File, Form
from fastapi.responses import StreamingResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, validator
from typing import List, Optional, Dict, Any
import uuid
from datetime import datetime
from enum import Enum
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from io import BytesIO
import pandas as pd

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# Create the main app without a prefix
app = FastAPI(title="Aplikasi Penilaian Guru", version="1.0.0")

# Create a router with the /api prefix
api_router = APIRouter(prefix="/api")

# Enums
class GenderEnum(str, Enum):
    LAKI_LAKI = "Laki-laki"
    PEREMPUAN = "Perempuan"

class StatusEnum(str, Enum):
    AKTIF = "Aktif"
    TIDAK_AKTIF = "Tidak Aktif"

# Models
class Student(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    nama: str = Field(..., min_length=1)
    nis: str = Field(..., min_length=1)
    kelas: str = Field(..., min_length=1)
    jenis_kelamin: GenderEnum
    status: StatusEnum = StatusEnum.AKTIF
    created_at: datetime = Field(default_factory=datetime.utcnow)
    updated_at: datetime = Field(default_factory=datetime.utcnow)

    @validator('nama')
    def validate_nama(cls, v):
        return v.strip().title()
    
    @validator('nis', 'kelas')
    def validate_nis_kelas(cls, v):
        return v.strip().upper()

class StudentCreate(BaseModel):
    nama: str = Field(..., min_length=1)
    nis: str = Field(..., min_length=1)
    kelas: str = Field(..., min_length=1)
    jenis_kelamin: GenderEnum
    status: StatusEnum = StatusEnum.AKTIF

class StudentUpdate(BaseModel):
    nama: Optional[str] = None
    nis: Optional[str] = None
    kelas: Optional[str] = None
    jenis_kelamin: Optional[GenderEnum] = None
    status: Optional[StatusEnum] = None

class Subject(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    nama_mata_pelajaran: str = Field(..., min_length=1)
    created_at: datetime = Field(default_factory=datetime.utcnow)
    updated_at: datetime = Field(default_factory=datetime.utcnow)

    @validator('nama_mata_pelajaran')
    def validate_nama_mata_pelajaran(cls, v):
        return v.strip().title()

class SubjectCreate(BaseModel):
    nama_mata_pelajaran: str = Field(..., min_length=1)

class LearningObjective(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    tujuan_pembelajaran: str = Field(..., min_length=1)
    created_at: datetime = Field(default_factory=datetime.utcnow)
    updated_at: datetime = Field(default_factory=datetime.utcnow)

    @validator('tujuan_pembelajaran')
    def validate_tujuan_pembelajaran(cls, v):
        return v.strip()

class LearningObjectiveCreate(BaseModel):
    tujuan_pembelajaran: str = Field(..., min_length=1)

class SubjectClassObjective(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    subject_id: str
    kelas: str
    learning_objective_ids: List[str]
    created_at: datetime = Field(default_factory=datetime.utcnow)
    updated_at: datetime = Field(default_factory=datetime.utcnow)

class SubjectClassObjectiveCreate(BaseModel):
    subject_id: str
    kelas: str
    learning_objective_ids: List[str]

class Grade(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    student_id: str
    subject_id: str
    kelas: str
    learning_objective_id: str
    nilai: float = Field(..., ge=0, le=100)
    created_at: datetime = Field(default_factory=datetime.utcnow)
    updated_at: datetime = Field(default_factory=datetime.utcnow)

class GradeCreate(BaseModel):
    student_id: str
    subject_id: str
    kelas: str
    learning_objective_id: str
    nilai: float = Field(..., ge=0, le=100)

class GradeUpdate(BaseModel):
    nilai: float = Field(..., ge=0, le=100)

# Helper functions
async def get_student_by_nis(nis: str):
    student = await db.students.find_one({"nis": nis.upper()})
    return student

async def get_subject_class_objective(subject_id: str, kelas: str):
    sco = await db.subject_class_objectives.find_one({
        "subject_id": subject_id,
        "kelas": kelas.upper()
    })
    return sco

# Student Management Endpoints
@api_router.post("/students", response_model=Student)
async def create_student(student: StudentCreate):
    # Check for duplicate NIS
    existing_student = await get_student_by_nis(student.nis)
    if existing_student:
        raise HTTPException(status_code=400, detail=f"Siswa dengan NIS {student.nis} sudah ada")
    
    student_dict = student.dict()
    student_obj = Student(**student_dict)
    await db.students.insert_one(student_obj.dict())
    return student_obj

@api_router.get("/students", response_model=List[Student])
async def get_students(search: Optional[str] = None, kelas: Optional[str] = None):
    query = {}
    if search:
        query["$or"] = [
            {"nama": {"$regex": search, "$options": "i"}},
            {"nis": {"$regex": search, "$options": "i"}}
        ]
    if kelas:
        query["kelas"] = kelas.upper()
    
    students = await db.students.find(query).sort("nama", 1).to_list(1000)
    return [Student(**{k: v for k, v in student.items() if k != "_id"}) for student in students]

@api_router.get("/students/{student_id}", response_model=Student)
async def get_student(student_id: str):
    student = await db.students.find_one({"id": student_id})
    if not student:
        raise HTTPException(status_code=404, detail="Siswa tidak ditemukan")
    return Student(**student)

@api_router.put("/students/{student_id}", response_model=Student)
async def update_student(student_id: str, student_update: StudentUpdate):
    student = await db.students.find_one({"id": student_id})
    if not student:
        raise HTTPException(status_code=404, detail="Siswa tidak ditemukan")
    
    update_data = student_update.dict(exclude_unset=True)
    if update_data:
        # Check for duplicate NIS if NIS is being updated
        if "nis" in update_data:
            existing_student = await get_student_by_nis(update_data["nis"])
            if existing_student and existing_student["id"] != student_id:
                raise HTTPException(status_code=400, detail=f"Siswa dengan NIS {update_data['nis']} sudah ada")
        
        update_data["updated_at"] = datetime.utcnow()
        await db.students.update_one({"id": student_id}, {"$set": update_data})
        
    updated_student = await db.students.find_one({"id": student_id})
    return Student(**updated_student)

@api_router.delete("/students/{student_id}")
async def delete_student(student_id: str):
    result = await db.students.delete_one({"id": student_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Siswa tidak ditemukan")
    
    # Delete related grades
    await db.grades.delete_many({"student_id": student_id})
    return {"message": "Siswa berhasil dihapus"}

@api_router.delete("/students")
async def delete_all_students():
    await db.students.delete_many({})
    await db.grades.delete_many({})
    return {"message": "Semua data siswa berhasil dihapus"}

@api_router.get("/students/classes/list")
async def get_classes():
    classes = await db.students.distinct("kelas")
    return sorted(classes)

# Subject Management Endpoints
@api_router.post("/subjects", response_model=Subject)
async def create_subject(subject: SubjectCreate):
    # Check for duplicate subject name
    existing_subject = await db.subjects.find_one({"nama_mata_pelajaran": subject.nama_mata_pelajaran.strip().title()})
    if existing_subject:
        raise HTTPException(status_code=400, detail=f"Mata pelajaran {subject.nama_mata_pelajaran} sudah ada")
    
    subject_dict = subject.dict()
    subject_obj = Subject(**subject_dict)
    await db.subjects.insert_one(subject_obj.dict())
    return subject_obj

@api_router.get("/subjects", response_model=List[Subject])
async def get_subjects():
    subjects = await db.subjects.find().sort("nama_mata_pelajaran", 1).to_list(1000)
    return [Subject(**{k: v for k, v in subject.items() if k != "_id"}) for subject in subjects]

@api_router.put("/subjects/{subject_id}", response_model=Subject)
async def update_subject(subject_id: str, subject_update: SubjectCreate):
    subject = await db.subjects.find_one({"id": subject_id})
    if not subject:
        raise HTTPException(status_code=404, detail="Mata pelajaran tidak ditemukan")
    
    # Check for duplicate subject name
    existing_subject = await db.subjects.find_one({
        "nama_mata_pelajaran": subject_update.nama_mata_pelajaran.strip().title(),
        "id": {"$ne": subject_id}
    })
    if existing_subject:
        raise HTTPException(status_code=400, detail=f"Mata pelajaran {subject_update.nama_mata_pelajaran} sudah ada")
    
    update_data = subject_update.dict()
    update_data["updated_at"] = datetime.utcnow()
    await db.subjects.update_one({"id": subject_id}, {"$set": update_data})
    
    updated_subject = await db.subjects.find_one({"id": subject_id})
    return Subject(**updated_subject)

@api_router.delete("/subjects/{subject_id}")
async def delete_subject(subject_id: str):
    result = await db.subjects.delete_one({"id": subject_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Mata pelajaran tidak ditemukan")
    
    # Delete related data
    await db.subject_class_objectives.delete_many({"subject_id": subject_id})
    await db.grades.delete_many({"subject_id": subject_id})
    return {"message": "Mata pelajaran berhasil dihapus"}

# Learning Objective Management Endpoints
@api_router.post("/learning-objectives", response_model=LearningObjective)
async def create_learning_objective(objective: LearningObjectiveCreate):
    objective_dict = objective.dict()
    objective_obj = LearningObjective(**objective_dict)
    await db.learning_objectives.insert_one(objective_obj.dict())
    return objective_obj

@api_router.get("/learning-objectives", response_model=List[LearningObjective])
async def get_learning_objectives():
    objectives = await db.learning_objectives.find().sort("tujuan_pembelajaran", 1).to_list(1000)
    return [LearningObjective(**{k: v for k, v in objective.items() if k != "_id"}) for objective in objectives]

@api_router.put("/learning-objectives/{objective_id}", response_model=LearningObjective)
async def update_learning_objective(objective_id: str, objective_update: LearningObjectiveCreate):
    objective = await db.learning_objectives.find_one({"id": objective_id})
    if not objective:
        raise HTTPException(status_code=404, detail="Tujuan pembelajaran tidak ditemukan")
    
    update_data = objective_update.dict()
    update_data["updated_at"] = datetime.utcnow()
    await db.learning_objectives.update_one({"id": objective_id}, {"$set": update_data})
    
    updated_objective = await db.learning_objectives.find_one({"id": objective_id})
    return LearningObjective(**updated_objective)

@api_router.delete("/learning-objectives/{objective_id}")
async def delete_learning_objective(objective_id: str):
    result = await db.learning_objectives.delete_one({"id": objective_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Tujuan pembelajaran tidak ditemukan")
    
    # Delete related data
    await db.subject_class_objectives.delete_many({"learning_objective_ids": objective_id})
    await db.grades.delete_many({"learning_objective_id": objective_id})
    return {"message": "Tujuan pembelajaran berhasil dihapus"}

# Subject Class Objective Management
@api_router.post("/subject-class-objectives", response_model=SubjectClassObjective)
async def create_subject_class_objective(sco: SubjectClassObjectiveCreate):
    # Check if combination already exists
    existing_sco = await get_subject_class_objective(sco.subject_id, sco.kelas)
    if existing_sco:
        raise HTTPException(status_code=400, detail="Konfigurasi mata pelajaran untuk kelas ini sudah ada")
    
    sco_dict = sco.dict()
    sco_dict["kelas"] = sco.kelas.upper()
    sco_obj = SubjectClassObjective(**sco_dict)
    await db.subject_class_objectives.insert_one(sco_obj.dict())
    return sco_obj

@api_router.get("/subject-class-objectives")
async def get_subject_class_objectives():
    scos = await db.subject_class_objectives.find().to_list(1000)
    result = []
    
    for sco in scos:
        # Get subject info
        subject = await db.subjects.find_one({"id": sco["subject_id"]})
        
        # Get learning objectives info
        objectives = []
        for obj_id in sco["learning_objective_ids"]:
            obj = await db.learning_objectives.find_one({"id": obj_id})
            if obj:
                # Remove MongoDB _id field to prevent serialization issues
                obj_clean = {k: v for k, v in obj.items() if k != "_id"}
                objectives.append(obj_clean)
        
        # Clean subject data
        subject_clean = {k: v for k, v in subject.items() if k != "_id"} if subject else None
        
        result.append({
            "id": sco["id"],
            "subject": subject_clean,
            "kelas": sco["kelas"],
            "learning_objectives": objectives,
            "created_at": sco["created_at"]
        })
    
    return result

@api_router.put("/subject-class-objectives/{sco_id}")
async def update_subject_class_objective(sco_id: str, sco_update: SubjectClassObjectiveCreate):
    sco = await db.subject_class_objectives.find_one({"id": sco_id})
    if not sco:
        raise HTTPException(status_code=404, detail="Konfigurasi tidak ditemukan")
    
    # Check if new combination already exists (excluding current record)
    existing_sco = await db.subject_class_objectives.find_one({
        "subject_id": sco_update.subject_id,
        "kelas": sco_update.kelas.upper(),
        "id": {"$ne": sco_id}
    })
    if existing_sco:
        raise HTTPException(status_code=400, detail="Konfigurasi mata pelajaran untuk kelas ini sudah ada")
    
    update_data = sco_update.dict()
    update_data["kelas"] = sco_update.kelas.upper()
    update_data["updated_at"] = datetime.utcnow()
    await db.subject_class_objectives.update_one({"id": sco_id}, {"$set": update_data})
    
    return {"message": "Konfigurasi berhasil diperbarui"}

@api_router.delete("/subject-class-objectives/{sco_id}")
async def delete_subject_class_objective(sco_id: str):
    result = await db.subject_class_objectives.delete_one({"id": sco_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Konfigurasi tidak ditemukan")
    
    return {"message": "Konfigurasi berhasil dihapus"}

# Grade Management Endpoints
@api_router.get("/grades/objectives/{subject_id}/{kelas}")
async def get_objectives_by_subject_class(subject_id: str, kelas: str):
    sco = await get_subject_class_objective(subject_id, kelas)
    if not sco:
        return []
    
    objectives = []
    for obj_id in sco["learning_objective_ids"]:
        obj = await db.learning_objectives.find_one({"id": obj_id})
        if obj:
            # Remove MongoDB _id field to prevent serialization issues
            obj_clean = {k: v for k, v in obj.items() if k != "_id"}
            objectives.append(LearningObjective(**obj_clean))
    
    return objectives

@api_router.get("/grades/{subject_id}/{kelas}/{objective_id}")
async def get_grades_by_criteria(subject_id: str, kelas: str, objective_id: str):
    # Get students in the class
    students = await db.students.find({"kelas": kelas.upper()}).sort("nama", 1).to_list(1000)
    
    result = []
    for student in students:
        # Get grade for this student and objective
        grade = await db.grades.find_one({
            "student_id": student["id"],
            "subject_id": subject_id,
            "kelas": kelas.upper(),
            "learning_objective_id": objective_id
        })
        
        # Clean student data
        student_clean = {k: v for k, v in student.items() if k != "_id"}
        grade_clean = {k: v for k, v in grade.items() if k != "_id"} if grade else None
        
        result.append({
            "student": Student(**student_clean),
            "grade": Grade(**grade_clean) if grade_clean else None
        })
    
    return result

@api_router.post("/grades", response_model=Grade)
async def create_or_update_grade(grade_data: GradeCreate):
    # Check if grade already exists
    existing_grade = await db.grades.find_one({
        "student_id": grade_data.student_id,
        "subject_id": grade_data.subject_id,
        "kelas": grade_data.kelas.upper(),
        "learning_objective_id": grade_data.learning_objective_id
    })
    
    if existing_grade:
        # Update existing grade
        update_data = {"nilai": grade_data.nilai, "updated_at": datetime.utcnow()}
        await db.grades.update_one({"id": existing_grade["id"]}, {"$set": update_data})
        updated_grade = await db.grades.find_one({"id": existing_grade["id"]})
        return Grade(**updated_grade)
    else:
        # Create new grade
        grade_dict = grade_data.dict()
        grade_dict["kelas"] = grade_data.kelas.upper()
        grade_obj = Grade(**grade_dict)
        await db.grades.insert_one(grade_obj.dict())
        return grade_obj

@api_router.get("/reports/grades/{kelas}")
async def get_class_grade_report(kelas: str):
    # Get all students in class
    students = await db.students.find({"kelas": kelas.upper()}).sort("nama", 1).to_list(1000)
    
    # Get all subject-class-objectives for this class
    scos = await db.subject_class_objectives.find({"kelas": kelas.upper()}).to_list(1000)
    
    result = []
    for student in students:
        # Clean student data
        student_clean = {k: v for k, v in student.items() if k != "_id"}
        
        student_data = {
            "student": Student(**student_clean),
            "grades": [],
            "average": 0
        }
        
        total_grades = 0
        grade_count = 0
        
        for sco in scos:
            # Get subject info
            subject = await db.subjects.find_one({"id": sco["subject_id"]})
            
            for obj_id in sco["learning_objective_ids"]:
                # Get objective info
                objective = await db.learning_objectives.find_one({"id": obj_id})
                
                # Get grade
                grade = await db.grades.find_one({
                    "student_id": student["id"],
                    "subject_id": sco["subject_id"],
                    "kelas": kelas.upper(),
                    "learning_objective_id": obj_id
                })
                
                grade_info = {
                    "subject": subject["nama_mata_pelajaran"] if subject else "",
                    "objective": objective["tujuan_pembelajaran"] if objective else "",
                    "nilai": grade["nilai"] if grade else None
                }
                
                student_data["grades"].append(grade_info)
                
                if grade:
                    total_grades += grade["nilai"]
                    grade_count += 1
        
        # Calculate average
        if grade_count > 0:
            student_data["average"] = round(total_grades / grade_count, 2)
        
        result.append(student_data)
    
    return result

# Excel Template and Import/Export Endpoints
@api_router.get("/students/template/download")
async def download_student_template():
    # Create Excel template
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Template Data Siswa"
    
    # Headers
    headers = ["Nama", "NIS", "Kelas", "Jenis Kelamin", "Status"]
    ws.append(headers)
    
    # Sample data
    ws.append(["Contoh Siswa", "12345", "X1", "Laki-laki", "Aktif"])
    
    # Style headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Add data validation comments
    ws["A3"] = "Petunjuk:"
    ws["A4"] = "- Nama: Nama lengkap siswa"
    ws["A5"] = "- NIS: Nomor Induk Siswa (bisa huruf/angka)"
    ws["A6"] = "- Kelas: Format X1, X2, XI1, XI2, XII1, XII2, dst"
    ws["A7"] = "- Jenis Kelamin: Laki-laki atau Perempuan"
    ws["A8"] = "- Status: Aktif atau Tidak Aktif"
    
    # Save to BytesIO
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    return StreamingResponse(
        BytesIO(excel_buffer.read()),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=template_data_siswa.xlsx"}
    )

@api_router.post("/students/import")
async def import_students_from_excel(file: UploadFile = File(...)):
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="File harus berformat Excel (.xlsx atau .xls)")
    
    try:
        # Read Excel file
        contents = await file.read()
        df = pd.read_excel(BytesIO(contents))
        
        # Validate columns
        required_columns = ["Nama", "NIS", "Kelas", "Jenis Kelamin", "Status"]
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            raise HTTPException(status_code=400, detail=f"Kolom yang hilang: {', '.join(missing_columns)}")
        
        imported_count = 0
        duplicate_count = 0
        error_rows = []
        
        for index, row in df.iterrows():
            try:
                # Skip empty rows
                if pd.isna(row['Nama']) or pd.isna(row['NIS']):
                    continue
                
                # Check for duplicate NIS
                existing_student = await get_student_by_nis(str(row['NIS']).strip())
                if existing_student:
                    duplicate_count += 1
                    continue
                
                # Validate gender
                gender = str(row['Jenis Kelamin']).strip()
                if gender not in ["Laki-laki", "Perempuan"]:
                    error_rows.append(f"Baris {index + 2}: Jenis kelamin tidak valid ({gender})")
                    continue
                
                # Validate status
                status = str(row['Status']).strip() if not pd.isna(row['Status']) else "Aktif"
                if status not in ["Aktif", "Tidak Aktif"]:
                    status = "Aktif"
                
                # Create student
                student_data = StudentCreate(
                    nama=str(row['Nama']).strip(),
                    nis=str(row['NIS']).strip(),
                    kelas=str(row['Kelas']).strip(),
                    jenis_kelamin=gender,
                    status=status
                )
                
                student_dict = student_data.dict()
                student_obj = Student(**student_dict)
                await db.students.insert_one(student_obj.dict())
                imported_count += 1
                
            except Exception as e:
                error_rows.append(f"Baris {index + 2}: {str(e)}")
        
        return {
            "message": f"Import selesai",
            "imported_count": imported_count,
            "duplicate_count": duplicate_count,
            "error_count": len(error_rows),
            "errors": error_rows[:10]  # Show first 10 errors
        }
        
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error membaca file Excel: {str(e)}")

@api_router.get("/reports/grades/{kelas}/export")
async def export_class_grades_to_excel(kelas: str):
    # Get grade report data
    report_data = await get_class_grade_report(kelas)
    
    if not report_data:
        raise HTTPException(status_code=404, detail="Tidak ada data untuk kelas ini")
    
    # Create Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Nilai Kelas {kelas}"
    
    # Create headers
    headers = ["No", "Nama", "NIS"]
    
    # Get unique subjects and objectives
    all_objectives = set()
    for student_data in report_data:
        for grade in student_data["grades"]:
            obj_key = f"{grade['subject']} - {grade['objective']}"
            all_objectives.add(obj_key)
    
    sorted_objectives = sorted(list(all_objectives))
    headers.extend(sorted_objectives)
    headers.append("Rata-rata")
    
    # Add headers to worksheet
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Add student data
    for row_num, student_data in enumerate(report_data, 2):
        # Basic info
        ws.cell(row=row_num, column=1, value=row_num - 1)
        ws.cell(row=row_num, column=2, value=student_data["student"].nama)
        ws.cell(row=row_num, column=3, value=student_data["student"].nis)
        
        # Grades
        for grade in student_data["grades"]:
            obj_key = f"{grade['subject']} - {grade['objective']}"
            if obj_key in sorted_objectives:
                col_num = headers.index(obj_key) + 1
                ws.cell(row=row_num, column=col_num, value=grade["nilai"] if grade["nilai"] is not None else "-")
        
        # Average
        avg_col = len(headers)
        ws.cell(row=row_num, column=avg_col, value=student_data["average"] if student_data["average"] > 0 else "-")
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min((max_length + 2) * 1.2, 50)  # Max width 50
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Add borders
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row in ws.iter_rows():
        for cell in row:
            cell.border = thin_border
            if cell.row > 1:  # Data rows
                cell.alignment = Alignment(horizontal="center")
    
    # Save to BytesIO
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    filename = f"nilai_kelas_{kelas.replace(' ', '_')}.xlsx"
    return StreamingResponse(
        BytesIO(excel_buffer.read()),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@api_router.get("/")
async def root():
    return {"message": "Aplikasi Penilaian Guru API", "version": "1.0.0"}

# Include the router in the main app
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()